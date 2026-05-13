from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def add_dataframe_sheet(workbook: Workbook, title: str, df: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title=title)
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 36)


def read_json(path: Path) -> dict:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


def generate_excel_report(output_dir: Path, report_path: Path) -> Path:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Executive Summary"

    repeat_metrics = read_json(output_dir / "repeat_purchase_metrics.json")
    ab_results = read_json(output_dir / "ab_test_results.json")

    summary.append(["Metric", "Value"])
    for key, value in repeat_metrics.items():
        summary.append([f"Repeat purchase {key}", value])

    conversion = ab_results.get("conversion_test", {})
    for key, value in conversion.items():
        summary.append([f"AB {key}", value])

    for cell in summary[1]:
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 22

    csv_sheets = {
        "Data Quality": "data_quality_summary.csv",
        "Segment Summary": "segment_summary.csv",
        "Repeat Predictions": "repeat_purchase_predictions.csv",
        "Repeat by State": "repeat_state_summary.csv",
        "Repeat by Category": "repeat_category_summary.csv",
        "Repeat Drivers": "repeat_feature_importance.csv",
        "Repeat Top K": "repeat_topk_summary.csv",
        "Sales Forecast": "sales_forecast.csv",
    }
    for sheet_name, filename in csv_sheets.items():
        path = output_dir / filename
        if path.exists():
            add_dataframe_sheet(workbook, sheet_name, pd.read_csv(path).head(5000))

    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)
    return report_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Excel management report.")
    parser.add_argument("--output-dir", type=Path, default=Path("data/output"))
    parser.add_argument("--report-path", type=Path, default=Path("data/output/ecommerce_growth_report.xlsx"))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    report_path = generate_excel_report(args.output_dir, args.report_path)
    print(f"Excel report written to {report_path}")


if __name__ == "__main__":
    main()
